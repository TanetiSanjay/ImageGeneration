import os
import socket
import threading
import pickle
import time
import torch
import datetime
import openpyxl
from openpyxl import load_workbook
from image_generator import generate_image

FORMAT = "utf-8"
IMAGE_FORMAT = "latin-1"
SIZE = 1024

if not os.path.exists('IP_PROMPT.xlsx'):
    workbook = openpyxl.Workbook()
    workbook.save('IP_PROMPT.xlsx')

clients = []
queue_lock = threading.Lock()


def handle_client(connection, address):

    global clients, queue_lock

    connection.send(f'[*] Your position in the queue is {clients.index(connection) + 1}'.encode(FORMAT))
    clients.append(clients.pop(clients.index(connection)))
    queue_lock.acquire()

    print(f'[+] Connected to {connection} at {address} successfully')

    prompt, count = str(connection.recv(SIZE).decode(FORMAT)).split(':')

    try:
        time.sleep(1)
        ip_address = address[0]
        client_name = socket.gethostbyaddr(ip_address)[0]

    except socket.herror:
        time.sleep(1)
        print(f"[+] Name of the client can't be retrieved")
        now = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        client_name = f'client_file_{now}'
        print(f"[+] Created new folder for client IP: {address[0]} on the name {client_name}")

    print(f"[+] IP address of the client is: {address[0]}\n[+] Name of the client is {client_name}")
    print(f"[+] The prompt for Image generation is: {prompt}")

    wb = load_workbook('IP_PROMPT.xlsx')
    ws = wb.active

    if not ws['A1'].value:
        ws['A1'] = 'CLIENT_IP'
        ws['B1'] = 'CLIENT_PORT'
        ws['C1'] = 'CLIENT_NAME'
        ws['D1'] = 'PROMPT'
        ws['E1'] = 'COUNT'
        row_num = 4

    else:
        row_num = ws.max_row + 1

    ws.cell(row=row_num, column=1, value=f'{address[0]}')
    ws.cell(row=row_num, column=2, value=f'{address[1]}')
    ws.cell(row=row_num, column=3, value=f'{socket.getfqdn(connection.getpeername()[0])}')
    ws.cell(row=row_num, column=4, value=f'{prompt}')
    ws.cell(row=row_num, column=5, value=f'{count}')

    wb.save('IP_PROMPT.xlsx')

    save_path = f'D:\\Coding\\Sem_2\\Sem_Project\\Code\\ImageGeneration\\BackEnd\\Images\\{client_name}'
    generate_image(prompt, 1, save_path)
    time.sleep(0.5)

    file_names = os.listdir(save_path)
    file_dir_bytes = pickle.dumps(file_names)
    connection.sendall(file_dir_bytes)

    for file in os.listdir(save_path):

        try:
            file_path = os.path.join(save_path, file)
            time.sleep(1)
            file_size = os.path.getsize(file_path)
            connection.send(f'{file_size}'.encode(FORMAT))

            f = open(file_path, 'rb')
            for image_bytes in f:
                connection.send(image_bytes)
            f.close()

            print(f'[+] \"{file}\" at \"{file_path}\" Has been sent to {address}')
            time.sleep(0.5)

        except ConnectionAbortedError as e:
            print(f'[+] Error: \"ConnectionAbortedError\"')
            print(f'[+] Client was disconnected due to: {e}')

    queue_lock.release()

    torch.cuda.empty_cache()
    print(f'[*] GPU memory has been cleared')
    time.sleep(0.5)

    connection.close()
    print(f"[+] Connection was closed ({address[0]}:{address[1]})\n")


class Server:
    def __init__(self, ip, port):
        print(f"[+] Server is starting..")
        time.sleep(1)
        self.ip = ip
        self.port = port
        self.server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.server.bind((self.ip, self.port))

    def connection(self):
        print(f"[+] Server is listening at ({self.ip}:{self.port})")

        while True:
            try:
                self.server.listen()
                connection, address = self.server.accept()
                clients.append(connection)

                thread = threading.Thread(target=handle_client, args=(connection, address))
                thread.start()
            except ConnectionResetError:
                print(f'[+] Client has disconnected due to error')


if __name__ == '__main__':
    IP = socket.gethostbyname(socket.gethostname())
    PORT = 5555
    server = Server(IP, PORT)
    server.connection()
